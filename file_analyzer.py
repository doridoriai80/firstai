import os
import PyPDF2
from docx import Document
import openpyxl
from PIL import Image
import io
import base64
from typing import Dict, List, Optional
import json

# OCR 기능을 선택적으로 import
try:
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("경고: pytesseract가 설치되지 않았습니다. 이미지 OCR 기능을 사용할 수 없습니다.")

class FileAnalyzer:
    """파일 분석 클래스"""
    
    def __init__(self, upload_dir: str = "uploads"):
        self.upload_dir = upload_dir
        self.supported_extensions = {
            '.txt': self._analyze_text,
            '.pdf': self._analyze_pdf,
            '.docx': self._analyze_docx,
            '.xlsx': self._analyze_excel,
            '.jpg': self._analyze_image,
            '.jpeg': self._analyze_image,
            '.png': self._analyze_image,
            '.gif': self._analyze_image
        }
        
        # 업로드 디렉토리 생성
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
    
    def analyze_file(self, file) -> Dict:
        """파일 분석 메인 함수"""
        try:
            # 파일 확장자 확인
            filename = file.filename
            file_extension = os.path.splitext(filename)[1].lower()
            
            if file_extension not in self.supported_extensions:
                return {
                    'success': False,
                    'error': f'지원하지 않는 파일 형식입니다: {file_extension}'
                }
            
            # 파일 저장
            file_path = os.path.join(self.upload_dir, filename)
            file.save(file_path)
            
            # 파일 분석
            analyzer_func = self.supported_extensions[file_extension]
            content = analyzer_func(file_path)
            
            # 파일 정보
            file_info = {
                'filename': filename,
                'file_size': os.path.getsize(file_path),
                'file_type': file_extension,
                'content': content
            }
            
            return {
                'success': True,
                'file_info': file_info
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'파일 분석 중 오류가 발생했습니다: {str(e)}'
            }
    
    def _analyze_text(self, file_path: str) -> str:
        """텍스트 파일 분석"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content
        except UnicodeDecodeError:
            # UTF-8로 읽기 실패시 다른 인코딩 시도
            with open(file_path, 'r', encoding='cp949') as f:
                content = f.read()
            return content
    
    def _analyze_pdf(self, file_path: str) -> str:
        """PDF 파일 분석"""
        content = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    content += f"\n--- 페이지 {page_num + 1} ---\n"
                    content += page.extract_text()
        except Exception as e:
            content = f"PDF 읽기 오류: {str(e)}"
        return content
    
    def _analyze_docx(self, file_path: str) -> str:
        """Word 문서 분석"""
        try:
            doc = Document(file_path)
            content = ""
            for paragraph in doc.paragraphs:
                content += paragraph.text + "\n"
            return content
        except Exception as e:
            return f"Word 문서 읽기 오류: {str(e)}"
    
    def _analyze_excel(self, file_path: str) -> str:
        """Excel 파일 분석"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            content = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"\n--- 시트: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        content += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
            
            return content
        except Exception as e:
            return f"Excel 파일 읽기 오류: {str(e)}"
    
    def _analyze_image(self, file_path: str) -> str:
        """이미지 파일 분석 (OCR)"""
        try:
            image = Image.open(file_path)
            
            if not OCR_AVAILABLE:
                return f"이미지 파일이 업로드되었습니다. (OCR 기능을 사용하려면 pytesseract를 설치하세요)\n파일명: {os.path.basename(file_path)}\n파일 크기: {os.path.getsize(file_path)} bytes"
            
            # OCR 수행
            try:
                text = pytesseract.image_to_string(image, lang='kor+eng')
                if text.strip():
                    return f"이미지에서 추출된 텍스트:\n{text}"
                else:
                    return "이미지에서 텍스트를 추출할 수 없습니다."
            except Exception as ocr_error:
                return f"OCR 처리 중 오류: {str(ocr_error)}\n이미지 파일이 업로드되었습니다."
                
        except Exception as e:
            return f"이미지 읽기 오류: {str(e)}"
    
    def get_file_summary(self, file_info: Dict) -> str:
        """파일 정보 요약"""
        filename = file_info['filename']
        file_size = file_info['file_size']
        file_type = file_info['file_type']
        content = file_info['content']
        
        # 파일 크기 변환
        if file_size < 1024:
            size_str = f"{file_size} B"
        elif file_size < 1024 * 1024:
            size_str = f"{file_size / 1024:.1f} KB"
        else:
            size_str = f"{file_size / (1024 * 1024):.1f} MB"
        
        # 내용 길이
        content_length = len(content)
        
        summary = f"""
📁 파일 정보:
• 파일명: {filename}
• 파일 크기: {size_str}
• 파일 형식: {file_type}
• 내용 길이: {content_length} 문자

📄 파일 내용 (처음 500자):
{content[:500]}{'...' if len(content) > 500 else ''}
"""
        return summary
    
    def cleanup_files(self, max_age_hours: int = 24):
        """오래된 파일 정리"""
        import time
        current_time = time.time()
        
        for filename in os.listdir(self.upload_dir):
            file_path = os.path.join(self.upload_dir, filename)
            file_age = current_time - os.path.getmtime(file_path)
            
            if file_age > max_age_hours * 3600:  # 시간을 초로 변환
                try:
                    os.remove(file_path)
                    print(f"삭제된 파일: {filename}")
                except Exception as e:
                    print(f"파일 삭제 실패: {filename}, 오류: {e}")
